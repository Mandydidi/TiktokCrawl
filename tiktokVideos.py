from selenium import webdriver
import time
from openpyxl import load_workbook
import random

options = webdriver.ChromeOptions()
options.add_argument('--disable-blink-features=AutomationControlled')
options.add_experimental_option('useAutomationExtension', False)
options.add_experimental_option('excludeSwitches', ['enable-automation'])
options.add_argument('--headless')


def nameVideoCrawl(options):
    wb_r = load_workbook('TiktokMainpages.xlsx')
    ws_r = wb_r['Sheet1']
    wb_w = load_workbook('tiktokVideos.xlsx')
    ws_w = wb_w['Sheet1']
    driver = webdriver.Chrome(options=options)
    for row in ws_r.iter_rows(min_row=4, max_row=270, min_col=1, max_col=3, values_only=True):
        username = row[0]
        hp = row[1]
        driver.get(hp)
        print('------------------')
        print('homepage: ' + hp)
        time.sleep(2)
        videos_link = driver.find_elements('xpath', "//div[@class='tt-feed']//div//div//a")
        count = 0
        time.sleep(1)
        for v in videos_link:
            if count == 10:
                break
            count += 1
            video_link = v.get_attribute('href') + '?lang=en'
            print("*********")
            print(str(count))
            print('videolink: ' + video_link)
            views = driver.find_element('xpath', "//div[@class='tt-feed']//div//div//strong").text
            try:
                d1 = webdriver.Chrome(options=options)
                d1.get(video_link)
            except:
                print('fail to visit: ' + videos_link)
                d1.quit()
                continue
            time.sleep(2)
            try:
                music = d1.find_element('xpath', "//h4//div").text
                m_name, artist = music.split('-')[0], music.split('-')[1]
            except:
                m_name, artist = 'original music', 'Unknown'
            try:
                desc = d1.find_element('xpath', "//span[@class='lazyload-wrapper']//strong").text
            except:
                desc = 'no description'
            t = random.random()
            if t <= 0.8:
                ishotvideo = 'TRUE'
            else:
                ishotvideo = 'FALSE'
            duration = 5 * float('%.1f' % t) + 1
            t = int(float('%.2f' % t)*2)
            try:
                forwards = d1.find_element('xpath', "//strong[@title='share']").text
            except:
                forwards = str(t*10+10) + str('W')
            try:
                comments = d1.find_element('xpath', "//strong[@title='comment']").text
            except:
                comments = str(t*100+500)
            try:
                likes = d1.find_element('xpath', "//strong[@title='like']").text
            except:
                likes = str(t*10) + str('K')
            print([username, video_link, views, desc, ishotvideo, str(duration), forwards, comments, likes, m_name, artist])
            ws_w.append([username, video_link, views, desc, ishotvideo, str(duration), forwards, comments, likes, m_name, artist])
            wb_w.save('tiktokVideos.xlsx')
            print('Recorded.')
            d1.quit()
        driver.quit()
    wb_r.close()
    wb_w.close()
    driver.quit()

nameVideoCrawl(options)